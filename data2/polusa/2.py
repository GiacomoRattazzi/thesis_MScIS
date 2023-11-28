import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Data from POLUSA
polusa_data = {
    'LEFT': ['AlterNet', 'HuffPost', 'Los Angeles Times', 'Mother Jones', 'NPR', 'PBS', 'Slate', 'The Economist', 'The Guardian', 'The Nation', 'The New York Times', 'ThinkProgress'],
    'CENTER': ['ABC News', 'CBS News', 'NBC News', 'Reuters', 'USA Today', 'Yahoo! News'],
    'RIGHT': ['Breitbart', 'CNS News', 'Fox News', 'National Review', 'The Blaze', 'The Daily Caller', 'The State', 'The Weekly Standard', 'Townhall'],
    'UNDEFINED': ['AOL', 'BBC', 'Bloomberg News', 'Chicago Tribune', 'CNN', 'Politico', 'MSNBC', 'Reason', 'The Wall Street Journal', 'The Washington Post']
}

# Read AllSides data from CSV
all_sides_df = pd.read_csv('news_article_counts_with_stance.csv')
all_sides_companies = all_sides_df['News Company'].tolist()

# Create DataFrame
df = pd.DataFrame(columns=['News Company', 'POLUSA', 'AllSides', 'POLUSA Stance', 'AllSides Stance'])

# Populate DataFrame
rows = []
for stance, companies in polusa_data.items():
    for company in companies:
        allsides_stance = all_sides_df[all_sides_df['News Company'] == company]['Political Stance (AllSides)'].values[0] if company in all_sides_companies else ''
        rows.append({
            'News Company': company,
            'POLUSA': 'Yes',
            'AllSides': 'Yes' if company in all_sides_companies else 'No',
            'POLUSA Stance': stance,
            'AllSides Stance': allsides_stance
        })

df = pd.concat([df, pd.DataFrame(rows)], ignore_index=True)

# Sort DataFrame by AllSides presence
df.sort_values(by='AllSides', ascending=False, inplace=True)

# Create Excel workbook
wb = Workbook()
ws = wb.active

# Add DataFrame to Excel
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# Format colors
green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

for row in ws.iter_rows(min_row=2, max_col=5, max_row=ws.max_row):
    # Color POLUSA column
    row[1].fill = green_fill if row[1].value == 'Yes' else red_fill
    # Color AllSides column
    row[2].fill = green_fill if row[2].value == 'Yes' else red_fill

# Save the workbook
wb.save('news_companies_stances.xlsx')